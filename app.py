from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from datetime import datetime
import os
import json
import openpyxl
from io import BytesIO

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mpl_league.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PLAYER_PHOTOS_FOLDER'] = 'player_photos'

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Database Models
class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    owner = db.Column(db.String(100), nullable=False)
    budget = db.Column(db.Integer, default=100000)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    players = db.relationship('AuctionedPlayer', backref='team', lazy=True, cascade='all, delete-orphan')

    def to_dict(self, include_players=False):
        data = {
            'id': self.id,
            'name': self.name,
            'owner': self.owner,
            'budget': self.budget,
            'total_spent': sum(p.price for p in self.players),
            'players_count': len(self.players),
            'available_budget': self.budget - sum(p.price for p in self.players)
        }
        
        if include_players:
            players_list = []
            for auctioned_player in self.players:
                player_data = auctioned_player.player_ref.to_dict()
                player_data['price'] = auctioned_player.price
                players_list.append(player_data)
            data['players'] = players_list
        
        return data


class Player(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    serial_number = db.Column(db.Integer, unique=True)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    photo_path = db.Column(db.String(200))
    is_available = db.Column(db.Boolean, default=True)
    auctioned_players = db.relationship('AuctionedPlayer', backref='player_ref', lazy=True)

    def to_dict(self):
        # Extract filename from photo_path
        photo_url = None
        if self.photo_path:
            photo_filename = os.path.basename(self.photo_path)
            photo_url = f'/photos/{photo_filename}'
        
        return {
            'id': self.id,
            'serial_number': self.serial_number,
            'name': self.name,
            'role': self.role,
            'photo_path': self.photo_path,
            'photo_url': photo_url,
            'is_available': self.is_available
        }


class AuctionedPlayer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    player_id = db.Column(db.Integer, db.ForeignKey('player.id'), nullable=False)
    price = db.Column(db.Integer, nullable=False)
    auctioned_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'team_id': self.team_id,
            'player': self.player_ref.to_dict(),
            'price': self.price,
            'auctioned_at': self.auctioned_at.isoformat()
        }


# Routes

@app.route('/photos/<filename>')
def get_photo(filename):
    try:
        photos_dir = app.config['PLAYER_PHOTOS_FOLDER']
        return send_from_directory(photos_dir, filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/bidding')
def bidding():
    return render_template('bidding.html')

@app.route('/api/teams', methods=['GET', 'POST'])
def teams():
    if request.method == 'POST':
        data = request.json
        team_name = data.get('name')
        team_owner = data.get('owner')
        
        if not team_name or not team_owner:
            return jsonify({'error': 'Team name and owner are required'}), 400
        
        if Team.query.filter_by(name=team_name).first():
            return jsonify({'error': 'Team already exists'}), 400
        
        team = Team(name=team_name, owner=team_owner)
        db.session.add(team)
        db.session.commit()
        return jsonify(team.to_dict()), 201
    
    teams_list = Team.query.all()
    return jsonify([t.to_dict(include_players=True) for t in teams_list])


@app.route('/api/teams/<int:team_id>', methods=['GET', 'DELETE'])
def team_detail(team_id):
    team = Team.query.get_or_404(team_id)
    
    if request.method == 'DELETE':
        db.session.delete(team)
        db.session.commit()
        return '', 204
    
    team_data = team.to_dict()
    team_data['players'] = [p.to_dict() for p in team.players]
    return jsonify(team_data)


@app.route('/api/players', methods=['GET', 'POST'])
def players():
    if request.method == 'POST':
        # Import players from Excel
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'No file provided'}), 400
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # serial_number and name
                    serial = row[0]
                    name = row[1]
                    role = row[2] if row[2] else 'Unknown'
                    
                    if not Player.query.filter_by(serial_number=serial).first():
                        player = Player(
                            serial_number=serial,
                            name=name,
                            role=role
                        )
                        db.session.add(player)
                        imported_count += 1
            
            db.session.commit()
            return jsonify({
                'message': f'{imported_count} players imported successfully',
                'count': imported_count
            }), 201
        except Exception as e:
            return jsonify({'error': str(e)}), 400
    
    # Get available players
    available_only = request.args.get('available', 'false').lower() == 'true'
    if available_only:
        players_list = Player.query.filter_by(is_available=True).all()
    else:
        players_list = Player.query.all()
    
    return jsonify([p.to_dict() for p in players_list])


@app.route('/api/players/<int:player_id>', methods=['GET', 'DELETE'])
def player_detail(player_id):
    player = Player.query.get_or_404(player_id)
    
    if request.method == 'DELETE':
        db.session.delete(player)
        db.session.commit()
        return '', 204
    
    return jsonify(player.to_dict())


@app.route('/api/auction', methods=['POST'])
def auction():
    data = request.json
    team_id = data.get('team_id')
    player_id = data.get('player_id')
    price = data.get('price', 1000)
    
    team = Team.query.get_or_404(team_id)
    player = Player.query.get_or_404(player_id)
    
    # Validation
    if not player.is_available:
        return jsonify({'error': 'Player already auctioned'}), 400
    
    available_budget = team.budget - sum(p.price for p in team.players)
    if price > available_budget:
        return jsonify({'error': f'Insufficient budget. Available: {available_budget}'}), 400
    
    if price < 1000:
        return jsonify({'error': 'Minimum bidding price is 1000'}), 400
    
    # Add player to team
    auctioned = AuctionedPlayer(
        team_id=team_id,
        player_id=player_id,
        price=price
    )
    player.is_available = False
    
    db.session.add(auctioned)
    db.session.commit()
    
    return jsonify({
        'message': f'{player.name} added to {team.name} for {price} points',
        'team': team.to_dict(),
        'auctioned_player': auctioned.to_dict()
    }), 201


@app.route('/api/auction/<int:auction_id>', methods=['DELETE'])
def remove_from_auction(auction_id):
    auctioned = AuctionedPlayer.query.get_or_404(auction_id)
    player = auctioned.player_ref
    
    player.is_available = True
    db.session.delete(auctioned)
    db.session.commit()
    
    return jsonify({'message': f'{player.name} removed from team'}), 200


@app.route('/api/auction/<int:team_id>/<int:player_id>', methods=['DELETE'])
def remove_player_from_team(team_id, player_id):
    """Remove a player from a team (by team_id and player_id)"""
    auctioned = AuctionedPlayer.query.filter_by(
        team_id=team_id,
        player_id=player_id
    ).first_or_404()
    
    player = auctioned.player_ref
    
    # Mark player as available again
    player.is_available = True
    db.session.delete(auctioned)
    db.session.commit()
    
    return jsonify({'message': f'{player.name} removed from team. Budget refunded!'}), 200


@app.route('/api/dashboard')
def dashboard():
    teams = Team.query.all()
    total_players = Player.query.count()
    available_players = Player.query.filter_by(is_available=True).count()
    auctioned_players = Player.query.filter_by(is_available=False).count()
    
    teams_data = []
    for team in teams:
        total_spent = sum(p.price for p in team.players)
        teams_data.append({
            'name': team.name,
            'spent': total_spent,
            'available': team.budget - total_spent,
            'players': len(team.players)
        })
    
    return jsonify({
        'total_players': total_players,
        'available_players': available_players,
        'auctioned_players': auctioned_players,
        'teams': teams_data
    })


@app.route('/api/export', methods=['GET'])
def export_data():
    teams = Team.query.all()
    
    wb = openpyxl.Workbook()
    
    # Teams sheet
    ws_teams = wb.active
    ws_teams.title = "Teams Summary"
    ws_teams.append(['Team Name', 'Budget', 'Spent', 'Available', 'Players Count'])
    
    for team in teams:
        spent = sum(p.price for p in team.players)
        ws_teams.append([
            team.name,
            team.budget,
            spent,
            team.budget - spent,
            len(team.players)
        ])
    
    # Team rosters
    for team in teams:
        ws = wb.create_sheet(title=team.name[:31])  # Excel sheet name limit
        ws.append(['Serial', 'Player Name', 'Role', 'Price'])
        
        for auctioned in team.players:
            player = auctioned.player_ref
            ws.append([
                player.serial_number,
                player.name,
                player.role,
                auctioned.price
            ])
    
    # Available players
    ws_available = wb.create_sheet(title="Available Players")
    ws_available.append(['Serial', 'Player Name', 'Role'])
    
    for player in Player.query.filter_by(is_available=True).all():
        ws_available.append([
            player.serial_number,
            player.name,
            player.role
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'MPL_League_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.before_request
def add_predefined_teams():
    if not hasattr(app, 'predefined_teams_added'):
        print("Adding predefined teams...")
        predefined_teams = [
            {"name": "subbi friends", "owner": "pushpak"},
            {"name": "RCB boys", "owner": "bharath"},
            {"name": "avi boys", "owner": "anvesh"},
            {"name": "coconut boys", "owner": "anil"},
            {"name": "nethravathi enterprises", "owner": "chethan"}
        ]

        for team in predefined_teams:
            print(f"Checking team: {team['name']}")
            if not Team.query.filter_by(name=team['name']).first():
                print(f"Adding team: {team['name']}")
                db.session.add(Team(name=team['name'], owner=team['owner']))
            else:
                print(f"Team already exists: {team['name']}")
        db.session.commit()
        print("Predefined teams added.")
        app.predefined_teams_added = True

if __name__ == "__main__":
    app.run(debug=True)

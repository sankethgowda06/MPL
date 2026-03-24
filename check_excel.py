import openpyxl

wb = openpyxl.load_workbook('Player_List.xlsx')
ws = wb.active

print("Checking Excel Data:\n")
print(f"{'Row':<5} {'Serial':<8} {'Player Name':<20} {'Role':<15}")
print("-" * 55)

for i in range(1, min(50, ws.max_row + 1)):
    serial = ws[f'A{i}'].value
    name = ws[f'B{i}'].value
    role = ws[f'C{i}'].value
    print(f"{i:<5} {str(serial):<8} {str(name):<20} {str(role):<15}")

print(f"\nTotal rows: {ws.max_row}")
